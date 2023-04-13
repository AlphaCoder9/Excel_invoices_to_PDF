import pandas as pd
import glob  # standard library
import openpyxl
from fpdf import FPDF
from pathlib import Path

filepaths = glob.glob('invoices/*.xlsx')

for filepath in filepaths:

    pdf = FPDF(orientation='P', unit='mm', format='A4')
    pdf.add_page()
    filename = Path(filepath).stem
    invoice_nr = filename.split('-')[0]
    Date = filename.split('-')[1]
    pdf.set_font(family='Times', size=16, style='B')
    pdf.cell(w=50, h=8, txt=f'Invoice nr.{invoice_nr}', ln=1)

    pdf.set_font(family='Times', size=16, style='B')
    pdf.cell(w=50, h=8, txt=f'Date: {Date}', ln=1)

    df = pd.read_excel(filepath, sheet_name='Sheet 1')
    for index, row in df.iterrows():
        pdf.set_font(family="Times", size=10)
        pdf.set_text_color(80, 80, 80)
        pdf.cell(w=30, h=8, txt=str(row["product_id"]), border=1)
        pdf.cell(w=70, h=8, txt=str(row["product_name"]), border=1)
        pdf.cell(w=30, h=8, txt=str(row['amount_purchased']), border=1)
        pdf.cell(w=30, h=8, txt=str(row['price_per_unit']), border=1)
        pdf.cell(w=30, h=8, txt=str(row['total_price']), border=1, ln=1)
    pdf.output(f'PDFs/{filename}.pdf')

